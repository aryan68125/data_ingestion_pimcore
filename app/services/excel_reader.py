import httpx
import orjson
from openpyxl import load_workbook

from app.utils.logger import LoggerFactory
from app.services.data_integrity_manager import ChunkIntegrityManager
from app.utils.logger_info_messages import ExcelInfoMessages
from app.utils.error_messages import ExcelErrorMessages

info_logger = LoggerFactory.get_info_logger()
error_logger = LoggerFactory.get_error_logger()
debug_logger = LoggerFactory.get_debug_logger()


class ExcelIngestionService:

    async def stream_and_push(self, ingestion_id: str, request):
        chunk = []
        chunk_number = 0
        total_records = 0
        
        info_logger.info(
            ExcelInfoMessages.STREAM_START.value.format(
                ingestion_id=ingestion_id
            )
        )
        info_logger.info(ExcelInfoMessages.WORKBOOK_LOAD_START.value)        
        
        wb = load_workbook(
            filename=request.file_path,
            read_only=True,
            data_only=True
        )
        info_logger.info(
            ExcelInfoMessages.WORKBOOK_LOADED.value
        )
            
        sheet = wb.active
        rows = sheet.iter_rows(values_only=True)

        header_row = next(rows, None)
        debug_logger.debug(
            f"Header row detected | header_row={header_row}"
        )

        if not header_row:
            error_logger.error(
                ExcelErrorMessages.EMPTY_HEADER.value.format(
                    ingestion_id=ingestion_id
                )
            )
            wb.close()
            return

        headers = [
            str(col).strip() if col is not None else f"column_{i}"
            for i, col in enumerate(header_row)
        ]
        debug_logger.debug(
            f"Headers parsed | headers={headers}"
        )
        async with httpx.AsyncClient(timeout=60) as client:
            for row in rows:
                
                if not any(row):
                    continue
                
                record = {
                    headers[i]: row[i]
                    for i in range(len(headers))
                }

                chunk.append(record)
                total_records += 1

                if request.chunk_size_by_records and len(chunk) >= request.chunk_size_by_records:
                    debug_logger.debug(
                        f"Chunk created | ingestion_id={ingestion_id} | "
                        f"chunk_number={chunk_number} | size={len(chunk)}"
                    )
                    await self._send_chunk(
                        client,
                        request.callback_url,
                        ingestion_id,
                        chunk_number,
                        chunk,
                        False
                    )
                    chunk_number += 1
                    chunk.clear()

            if chunk:
                debug_logger.debug(
                    f"Final chunk created | ingestion_id={ingestion_id} | "
                    f"chunk_number={chunk_number} | size={len(chunk)}"
                )
                await self._send_chunk(
                    client,
                    request.callback_url,
                    ingestion_id,
                    chunk_number,
                    chunk,
                    True
                )
            info_logger.info(
                ExcelInfoMessages.INGESTION_COMPLETED.value.format(
                    total_records=total_records
                )
            )
            await client.post(
                request.callback_url,
                json={
                    "ingestion_id": ingestion_id,
                    "status": "COMPLETED",
                    "total_records": total_records
                }
            )

        wb.close()

    async def _send_chunk(
        self,
        client,
        url,
        ingestion_id,
        chunk_number,
        records,
        is_last
    ):
        checksum = ChunkIntegrityManager.compute_checksum(records)
        chunk_id = ChunkIntegrityManager.build_chunk_id(
                ingestion_id, chunk_number
            )
   
        payload = {
            "ingestion_id": ingestion_id,
            "chunk_number": chunk_number,
            "chunk_id": chunk_id,
            "checksum": checksum,
            "records": records,
            "is_last": is_last
        }


        for attempt in range(3):
            try:
                debug_logger.debug(
                    f"Sending chunk | chunk_number={chunk_number} | "
                    f"attempt={attempt + 1} | records={len(records)}"
                )
                resp = await client.post(
                    url,
                    content=orjson.dumps(payload),
                    headers={"Content-Type": "application/json"}
                )

                ack_response = resp.json()
                debug_logger.debug(
                    f"Pimcore callback response | response={ack_response}"
                )
                ack = ack_response.get("ack")

                if ack is not True:
                    error = ack_response.get("error")
                    error_logger.error(
                        ExcelErrorMessages.CHUNK_REJECTED.value.format(
                            chunk_number=chunk_number,
                            reason=error
                        )
                    )
                    raise Exception(
                        f"Chunk {chunk_number} rejected: {error}"
                    )

                return

            except Exception as e:
                error_logger.error(
                    ExcelErrorMessages.CHUNK_PUSH_FAILED.value.format(
                        chunk_number=chunk_number,
                        attempt=attempt + 1,
                        error=str(e)
                    ),
                    exc_info=True
                )
                if attempt == 2:
                    raise
